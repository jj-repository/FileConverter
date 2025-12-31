from pathlib import Path
from typing import Dict, Any
import pandas as pd
from io import BytesIO
import asyncio

# Excel support
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ODS support
try:
    from odf import opendocument
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False

from app.services.base_converter import BaseConverter
from app.config import settings


class SpreadsheetConverter(BaseConverter):
    """Spreadsheet conversion service for Excel, ODS, CSV, TSV"""

    def __init__(self, websocket_manager=None):
        super().__init__(websocket_manager)
        self.supported_formats = {
            "input": list(settings.SPREADSHEET_FORMATS),
            "output": list(settings.SPREADSHEET_FORMATS),
        }

    async def get_supported_formats(self) -> Dict[str, list]:
        """Get supported spreadsheet formats"""
        return self.supported_formats

    async def convert(
        self,
        input_path: Path,
        output_format: str,
        options: Dict[str, Any],
        session_id: str
    ) -> Path:
        """
        Convert spreadsheet to target format

        Args:
            input_path: Path to input spreadsheet
            output_format: Target format (xlsx, xls, ods, csv, tsv)
            options: Conversion options
                - sheet_name: str (sheet to convert, default: first sheet)
                - include_all_sheets: bool (for multi-sheet formats, default: False)
                - delimiter: str (for CSV/TSV, default: auto)
                - encoding: str (default: utf-8)

        Returns:
            Path to converted spreadsheet
        """
        await self.send_progress(session_id, 0, "converting", "Starting spreadsheet conversion")

        # Validate format
        input_format = input_path.suffix.lower().lstrip('.')
        if not self.validate_format(input_format, output_format, self.supported_formats):
            raise ValueError(f"Unsupported conversion: {input_format} to {output_format}")

        # Generate output path
        output_path = settings.UPLOAD_DIR / f"{input_path.stem}_converted.{output_format}"

        # Get options
        sheet_name = options.get('sheet_name')
        include_all_sheets = options.get('include_all_sheets', False)
        encoding = options.get('encoding', 'utf-8')
        delimiter = options.get('delimiter')

        await self.send_progress(session_id, 20, "converting", "Reading input spreadsheet")

        try:
            # Read input file (wrap pandas I/O in thread pool)
            if input_format in ['xlsx', 'xls']:
                if not OPENPYXL_AVAILABLE:
                    raise ValueError("Excel support not available. Install openpyxl.")
                df = await self._read_excel(input_path, sheet_name)
            elif input_format == 'ods':
                if not ODF_AVAILABLE:
                    raise ValueError("ODS support not available. Install odfpy.")
                df = await self._read_ods(input_path, sheet_name)
            elif input_format == 'csv':
                # Auto-detect delimiter if not provided
                if not delimiter:
                    delimiter = ','
                df = await asyncio.to_thread(pd.read_csv, input_path, encoding=encoding, delimiter=delimiter)
            elif input_format == 'tsv':
                df = await asyncio.to_thread(pd.read_csv, input_path, encoding=encoding, delimiter='\t')
            else:
                raise ValueError(f"Unsupported input format: {input_format}")

            await self.send_progress(session_id, 60, "converting", "Converting spreadsheet format")

            # Write output file (wrap pandas I/O in thread pool)
            if output_format in ['xlsx', 'xls']:
                if not OPENPYXL_AVAILABLE:
                    raise ValueError("Excel support not available. Install openpyxl.")
                # openpyxl only writes XLSX, not XLS
                if output_format == 'xls':
                    raise ValueError("XLS output not supported. Use XLSX instead. (XLS reading is supported)")
                await asyncio.to_thread(df.to_excel, output_path, index=False, engine='openpyxl')
            elif output_format == 'ods':
                if not ODF_AVAILABLE:
                    raise ValueError("ODS support not available. Install odfpy.")
                await self._write_ods(df, output_path)
            elif output_format == 'csv':
                # Use delimiter from options or default to comma
                if not delimiter:
                    delimiter = ','
                await asyncio.to_thread(df.to_csv, output_path, index=False, encoding=encoding, sep=delimiter)
            elif output_format == 'tsv':
                await asyncio.to_thread(df.to_csv, output_path, index=False, encoding=encoding, sep='\t')
            else:
                raise ValueError(f"Unsupported output format: {output_format}")

            await self.send_progress(session_id, 100, "completed", "Spreadsheet conversion completed")

            return output_path

        except Exception as e:
            await self.send_progress(session_id, 0, "failed", f"Conversion failed: {str(e)}")
            raise

    async def _read_excel(self, file_path: Path, sheet_name: str = None) -> pd.DataFrame:
        """Read Excel file (XLSX or XLS)"""
        if sheet_name:
            return await asyncio.to_thread(pd.read_excel, file_path, sheet_name=sheet_name, engine='openpyxl')
        else:
            # Read first sheet by default
            return await asyncio.to_thread(pd.read_excel, file_path, engine='openpyxl')

    async def _read_ods(self, file_path: Path, sheet_name: str = None) -> pd.DataFrame:
        """Read ODS file"""
        doc = await asyncio.to_thread(opendocument.load, str(file_path))

        # Get all sheets
        sheets = doc.spreadsheet.getElementsByType(Table)

        if not sheets:
            raise ValueError("No sheets found in ODS file")

        # Select sheet
        if sheet_name:
            # Find sheet by name
            target_sheet = None
            for sheet in sheets:
                if sheet.getAttribute('name') == sheet_name:
                    target_sheet = sheet
                    break
            if not target_sheet:
                raise ValueError(f"Sheet '{sheet_name}' not found")
        else:
            # Use first sheet
            target_sheet = sheets[0]

        # Extract data
        data = []
        rows = target_sheet.getElementsByType(TableRow)

        for row in rows:
            row_data = []
            cells = row.getElementsByType(TableCell)
            for cell in cells:
                # Get cell value
                paragraphs = cell.getElementsByType(P)
                if paragraphs:
                    cell_value = ''.join([p.firstChild.data if p.firstChild else '' for p in paragraphs])
                else:
                    cell_value = ''
                row_data.append(cell_value)
            data.append(row_data)

        # Convert to DataFrame
        if data:
            # First row as headers
            if len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
            else:
                df = pd.DataFrame(data)
        else:
            df = pd.DataFrame()

        return df

    async def _write_ods(self, df: pd.DataFrame, output_path: Path):
        """Write DataFrame to ODS file"""
        # Create new ODS document
        doc = opendocument.OpenDocumentSpreadsheet()

        # Create table
        table = Table(name="Sheet1")

        # Add header row
        header_row = TableRow()
        for col in df.columns:
            cell = TableCell()
            p = P(text=str(col))
            cell.addElement(p)
            header_row.addElement(cell)
        table.addElement(header_row)

        # Add data rows
        for _, row in df.iterrows():
            table_row = TableRow()
            for value in row:
                cell = TableCell()
                p = P(text=str(value) if pd.notna(value) else '')
                cell.addElement(p)
                table_row.addElement(cell)
            table.addElement(table_row)

        doc.spreadsheet.addElement(table)
        await asyncio.to_thread(doc.save, str(output_path))

    async def get_spreadsheet_info(self, file_path: Path) -> Dict[str, Any]:
        """Extract metadata from spreadsheet file"""
        try:
            input_format = file_path.suffix.lower().lstrip('.')

            info = {
                "format": input_format,
                "size": file_path.stat().st_size,
            }

            if input_format in ['xlsx', 'xls']:
                if OPENPYXL_AVAILABLE:
                    # Load workbook (wrap in thread pool)
                    wb = await asyncio.to_thread(openpyxl.load_workbook, file_path, read_only=True, data_only=True)
                    info["sheets"] = len(wb.sheetnames)
                    info["sheet_names"] = wb.sheetnames

                    # Get info from first sheet
                    ws = wb.active
                    info["rows"] = ws.max_row
                    info["columns"] = ws.max_column
                    wb.close()

            elif input_format == 'ods':
                if ODF_AVAILABLE:
                    doc = await asyncio.to_thread(opendocument.load, str(file_path))
                    sheets = doc.spreadsheet.getElementsByType(Table)
                    info["sheets"] = len(sheets)
                    info["sheet_names"] = [sheet.getAttribute('name') for sheet in sheets]

                    if sheets:
                        rows = sheets[0].getElementsByType(TableRow)
                        info["rows"] = len(rows)

            elif input_format == 'csv':
                df = await asyncio.to_thread(pd.read_csv, file_path, nrows=1)
                full_df = await asyncio.to_thread(pd.read_csv, file_path)
                info["rows"] = len(full_df)
                info["columns"] = len(df.columns)
                info["column_names"] = df.columns.tolist()

            elif input_format == 'tsv':
                df = await asyncio.to_thread(pd.read_csv, file_path, delimiter='\t', nrows=1)
                full_df = await asyncio.to_thread(pd.read_csv, file_path, delimiter='\t')
                info["rows"] = len(full_df)
                info["columns"] = len(df.columns)
                info["column_names"] = df.columns.tolist()

            return info

        except Exception as e:
            return {"error": str(e)}
